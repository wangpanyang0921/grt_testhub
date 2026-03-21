      
#!/usr/bin/env python
# _*_ coding:utf-8 _*_  
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from xmind2testcase.utils import get_xmind_testcase_list, get_absolute_path

"""
Convert XMind file to Excel XLSX file

This module provides functionality to convert XMind testcase files to Excel XLSX format,
which is widely used in various test management systems and tools.
"""

def xmind_to_xlsx_file(xmind_file, case_owner='王盼阳'):
    """Convert XMind file to a XLSX file"""
    xmind_file = get_absolute_path(xmind_file)
    logging.info('Start converting XMind file(%s) to XLSX file...', xmind_file)
    testcases = get_xmind_testcase_list(xmind_file)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 设置表头，完全与CSV文件相同
    # fileheader = ["用例编号", "模块", "用例名称", "前置条件", "测试步骤1", "预期结果1", "执行方式", "用例级别", "用例类型", "测试步骤模式"]
    fileheader = ["标题", "目录","前置条件", "步骤描述", "预期结果", "负责人", "优先级", "类型", "标签", "预计工时汇总", "实际工时汇总"]
    ws.append(fileheader)
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充测试用例数据 - 使用与CSV完全相同的生成逻辑
    for index, testcase in enumerate(testcases, start=2):
        row = gen_a_testcase_row(testcase, xmind_file[:-6], case_owner)
        ws.append(row)
        
        # 设置单元格边框和对齐方式
        for cell in ws[index]:
            cell.border = thin_border
            # 所有单元格内容顶端对齐
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # 对于有大量文本的单元格，确保自动换行
            if cell.column_letter in ['E', 'F']:  # E: 测试步骤, F: 预期结果
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # 自动调整列宽和行高
    adjust_column_widths(ws)
    adjust_row_heights(ws)
    
    # 保存文件
    xlsx_file = xmind_file[:-6] + '.xlsx'
    if os.path.exists(xlsx_file):
        os.remove(xlsx_file)
        
    wb.save(xlsx_file)
    logging.info('Convert XMind file(%s) to a XLSX file(%s) successfully!', xmind_file, xlsx_file)
    
    return xlsx_file

def gen_a_testcase_row(testcase_dict, xlsx_file=None, case_owner='王盼阳'):
    """Generate a testcase row for XLSX file - 与CSV完全相同的实现"""
    case_number = ''
    list = testcase_dict['status'].split(' ')
    str_result2 = "|".join(list[:-1])
    # 获取文件名（不含路径）并拼接到最前面
    file_name = os.path.basename(xlsx_file) if xlsx_file else ''
    case_module = (f"{file_name}|" if file_name else "") + \
                  gen_case_module(testcase_dict['suite']) + \
                  (f"|{str_result2}" if str_result2 else "")
    # case_module = gen_case_module(testcase_dict['suite'])+ (f"|{str_result2}" if str_result2 else "")
    case_title = testcase_dict['name']
    case_precontion = testcase_dict['preconditions']
    case_step, case_expected_result = gen_case_step_and_expected_result(testcase_dict['steps'])
    case_priority = gen_case_priority(testcase_dict['importance'])
    case_type = gen_case_type(testcase_dict['execution_type'])
    case_apply_phase = '文本模式'
    case_label = ''
    case_expecthours = ''
    case_actualhours = ''
    row = [case_title, case_module, case_precontion, case_step, case_expected_result, case_owner, case_priority, case_type, case_label, case_expecthours, case_actualhours]
    return row

def gen_case_module(module_name):
    """Generate module name - 与CSV完全相同的实现"""
    if module_name:
        module_name = module_name.replace('（', '(')
        module_name = module_name.replace('）', ')')
    else:
        module_name = '/'
    return module_name

def gen_case_step_and_expected_result(steps):
    """Generate test steps and expected results - 与CSV完全相同的实现"""
    case_step = ''
    case_expected_result = ''

    for step_dict in steps:
        case_step += str(step_dict['step_number']) + '. ' + step_dict['actions'].replace('\n', '').strip() + '\n'
        case_expected_result += str(step_dict['step_number']) + '. ' + \
            step_dict['expectedresults'].replace('\n', '').strip() + '\n' \
            if step_dict.get('expectedresults', '') else ''

    return case_step, case_expected_result

def gen_case_priority(priority):
    """Convert priority number to text - 与CSV完全相同的实现"""
    mapping = {1: 'P0', 2: 'P1', 3: 'P2', 4: 'P3'}
    if priority in mapping.keys():
        return mapping[priority]
    else:
        return 'Priority 0'

def gen_case_type(case_type):
    """Convert case type number to text - 与CSV完全相同的实现"""
    mapping = {1: '功能测试', 2: '自动'}
    if case_type in mapping.keys():
        return mapping[case_type]
    else:
        return '功能测试'

def adjust_column_widths(ws):
    """根据内容自适应调整列宽"""
    # 先获取每列的最大字符数，然后根据内容自适应调整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # 获取该列所有非空单元格的最大内容长度
        for cell in column:
            try:
                if cell.value:
                    # 考虑中文字符，每个中文字符按2个宽度计算
                    text_length = 0
                    for char in str(cell.value):
                        if ord(char) > 127:  # 中文字符
                            text_length += 2
                        else:
                            text_length += 1
                    max_length = max(max_length, text_length)
            except:
                pass
        
        # 设置列宽，最大不超过50，避免过宽
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def adjust_row_heights(ws):
    """根据内容自适应调整行高，并额外增加15的空间"""
    # 遍历所有行，计算合适的行高
    for row in ws.iter_rows():
        max_height = 15  # 默认行高
        
        for cell in row:
            try:
                if cell.value:
                    text = str(cell.value)
                    # 计算文本中的换行符数量，预估需要的行数
                    line_count = text.count('\n') + 1
                    
                    # 对于没有换行符的长文本，根据列宽估算需要的行数
                    if line_count == 1:
                        # 获取列宽
                        col_width = ws.column_dimensions[cell.column_letter].width
                        # 假设每个字符宽度为1.2，计算需要的行数
                        text_length = 0
                        for char in text:
                            if ord(char) > 127:  # 中文字符
                                text_length += 2
                            else:
                                text_length += 1
                        
                        if col_width > 0:
                            line_count = max(line_count, int(text_length / col_width) + 1)
                    
                    # 每行高度约15
                    estimated_height = line_count * 15
                    max_height = max(max_height, estimated_height)
            except:
                pass
        
        # 设置行高，在自适应计算的基础上增加15的额外空间，最大不超过220
        adjusted_height = min(max_height + 15, 220)
        ws.row_dimensions[row[0].row].height = adjusted_height

if __name__ == '__main__':
    # 测试代码
    xmind_file = '../docs/xmind_testcase_demo.xmind'
    xlsx_file = xmind_to_xlsx_file(xmind_file)
    print('Successfully converted XMind file to XLSX file:', xlsx_file)

    