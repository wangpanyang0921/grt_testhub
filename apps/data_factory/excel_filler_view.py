"""
Excel数据智能填充API
使用第三方库生成数据：
- faker: 姓名、手机号、邮箱、身份证、公司、地址、日期等
- jionlp: 省市区县数据
"""
import pandas as pd
import random
import string
import re
import io
import base64
import json
from datetime import datetime, timedelta
from typing import Dict, Any
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .tools.random_tools import RandomTools

# 导入第三方数据生成库
try:
    from faker import Faker
    HAS_FAKER = True
except ImportError:
    HAS_FAKER = False

try:
    import jionlp
    HAS_JIONLP = True
except ImportError:
    HAS_JIONLP = False

# 初始化 faker
if HAS_FAKER:
    fake = Faker('zh_CN')
else:
    fake = None


class TemplateInstructionParser:
    """模板说明解析器 - 解析填写须知中的字段约束
    
    支持单Sheet和多Sheet模板：
    - 单Sheet：所有字段说明在一个Sheet中
    - 多Sheet：所有Sheet的字段说明都在第一个Sheet中，格式为 "Sheet名-字段名: 约束"
    """
    
    @staticmethod
    def parse(instruction_text: str) -> Dict[str, Dict]:
        """
        解析模板说明文本，提取每个字段的约束
        
        返回格式:
        {
            '字段名': {
                'required': bool,  # 是否必填
                'max_length': int,  # 最大长度
                'options': list,  # 选项列表
                'description': str  # 原始描述
            }
        }
        """
        constraints = {}
        
        if not instruction_text or '填写须知' not in instruction_text:
            return constraints
        
        # 按行分割
        lines = instruction_text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('填写须知'):
                continue
            
            # 匹配格式: 数字、字段名: 约束描述
            # 支持多Sheet格式: "Sheet名-字段名: 约束" 或 "Sheet名_字段名: 约束"
            # 例如: "1、学员姓名: 必填,10个字以内" 或 "1、Sheet2-学员姓名: 必填,10个字以内"
            match = re.match(r'\d+、?\s*([^:]+)[:：]\s*(.+)', line)
            if match:
                field_name = match.group(1).strip()
                description = match.group(2).strip()
                
                constraint = {
                    'required': False,
                    'max_length': None,
                    'options': None,
                    'description': description
                }
                
                # 判断是否必填
                if '必填' in description:
                    constraint['required'] = True
                elif '非必填' in description or '选填' in description:
                    constraint['required'] = False
                
                # 提取长度限制
                length_match = re.search(r'(\d+)\s*个?字', description)
                if length_match:
                    constraint['max_length'] = int(length_match.group(1))
                
                # 提取字符限制
                char_match = re.search(r'(\d+)\s*个?字符', description)
                if char_match:
                    constraint['max_length'] = int(char_match.group(1))
                
                # 提取选项（格式: xxx/xxx/xxx 或 xxx、xxx、xxx 或 xxx,xxx,xxx）
                # 匹配"博士研究生/硕士研究生/本科" 或 "市区/郊区/乡镇" 等
                
                # 尝试匹配选项 - 优先匹配"xxx/xxx/xxx中一项"格式
                if '项' in description:
                    # 查找"xxx/xxx/xxx中一项"格式，提取"中一项"之前的部分
                    options_match = re.search(r'([^，。]+?)(?:中[一1]?项|之一)', description)
                    if options_match:
                        options_text = options_match.group(1).strip()
                        # 分割选项
                        for sep in ['/', '、', ',', '，']:
                            if sep in options_text:
                                options = [opt.strip() for opt in options_text.split(sep) if opt.strip()]
                                # 清理最后一个选项可能包含的"中"字
                                if options:
                                    options = [opt.rstrip('中一1项') for opt in options]
                                    options = [opt for opt in options if opt]
                                if len(options) >= 2:
                                    constraint['options'] = options
                                    break
                
                # 如果没有匹配到，尝试直接匹配斜杠分隔的内容
                if not constraint['options']:
                    slash_match = re.search(r'([^，。]{2,}?)/([^，。]{2,}?)', description)
                    if slash_match:
                        # 提取所有斜杠分隔的部分
                        parts = re.split(r'[/、,，]', description)
                        # 过滤掉描述性文字，保留选项
                        options = []
                        for part in parts:
                            part = part.strip()
                            # 排除常见的描述词
                            if part and not any(word in part for word in ['必填', '非必填', '以内', '支持', '需要', '系统', '标签', '字符', '文本']):
                                if len(part) <= 20:  # 选项通常不会太长
                                    options.append(part)
                        if len(options) >= 2:
                            constraint['options'] = options
                
                constraints[field_name] = constraint
        
        return constraints
    
    @staticmethod
    def parse_multi_sheet(instruction_text: str, first_sheet_name: str = 'Sheet1') -> Dict[str, Dict[str, Dict]]:
        """
        解析多Sheet模板的说明文本
        
        Args:
            instruction_text: 填写须知文本
            first_sheet_name: 第一个Sheet的实际名称（用于单Sheet模板或无前缀字段）
        
        返回格式:
        {
            'Sheet名': {
                '字段名': {
                    'required': bool,
                    'max_length': int,
                    'options': list,
                    'description': str
                }
            }
        }
        """
        sheet_constraints = {}
        
        if not instruction_text or '填写须知' not in instruction_text:
            return sheet_constraints
        
        # 按行分割
        lines = instruction_text.split('\n')
        current_sheet = None
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('填写须知'):
                continue
            
            # 检测Sheet标题行（例如 "Sheet1:" 或 "Sheet1"）
            sheet_title_match = re.match(r'^([\w\u4e00-\u9fa5]+)[：:]\s*$', line)
            if sheet_title_match:
                current_sheet = sheet_title_match.group(1).strip()
                if current_sheet not in sheet_constraints:
                    sheet_constraints[current_sheet] = {}
                continue
            
            # 匹配字段说明: 数字、字段名: 约束描述
            match = re.match(r'\d+、?\s*([^:]+)[:：]\s*(.+)', line)
            if match:
                field_name = match.group(1).strip()
                description = match.group(2).strip()
                
                # 检查是否包含Sheet前缀（如 "Sheet2-字段名" 或 "Sheet2_字段名"）
                sheet_prefix_match = re.match(r'^([\w\u4e00-\u9fa5]+)[-_](.+)$', field_name)
                if sheet_prefix_match:
                    sheet_name = sheet_prefix_match.group(1)
                    actual_field_name = sheet_prefix_match.group(2)
                    if sheet_name not in sheet_constraints:
                        sheet_constraints[sheet_name] = {}
                    target_sheet = sheet_name
                    target_field = actual_field_name
                else:
                    # 如果没有Sheet前缀，使用当前Sheet或第一个Sheet的实际名称
                    target_sheet = current_sheet or first_sheet_name
                    if target_sheet not in sheet_constraints:
                        sheet_constraints[target_sheet] = {}
                    target_field = field_name
                
                constraint = {
                    'required': False,
                    'max_length': None,
                    'options': None,
                    'description': description
                }
                
                # 判断是否必填
                if '必填' in description:
                    constraint['required'] = True
                elif '非必填' in description or '选填' in description:
                    constraint['required'] = False
                
                # 提取长度限制
                length_match = re.search(r'(\d+)\s*个?字', description)
                if length_match:
                    constraint['max_length'] = int(length_match.group(1))
                
                # 提取字符限制
                char_match = re.search(r'(\d+)\s*个?字符', description)
                if char_match:
                    constraint['max_length'] = int(char_match.group(1))
                
                # 提取选项
                if '项' in description:
                    # 查找"xxx/xxx/xxx中一项"格式，提取"中一项"之前的部分
                    options_match = re.search(r'([^，。]+?)(?:中[一1]?项|之一)', description)
                    if options_match:
                        options_text = options_match.group(1).strip()
                        for sep in ['/', '、', ',', '，']:
                            if sep in options_text:
                                options = [opt.strip() for opt in options_text.split(sep) if opt.strip()]
                                # 清理最后一个选项可能包含的"中"字
                                if options:
                                    options = [opt.rstrip('中一1项') for opt in options]
                                    options = [opt for opt in options if opt]
                                if len(options) >= 2:
                                    constraint['options'] = options
                                    break
                
                if not constraint['options']:
                    slash_match = re.search(r'([^，。]{2,}?)/([^，。]{2,}?)', description)
                    if slash_match:
                        parts = re.split(r'[/、,，]', description)
                        options = []
                        for part in parts:
                            part = part.strip()
                            if part and not any(word in part for word in ['必填', '非必填', '以内', '支持', '需要', '系统', '标签', '字符', '文本']):
                                if len(part) <= 20:
                                    options.append(part)
                        if len(options) >= 2:
                            constraint['options'] = options
                
                sheet_constraints[target_sheet][target_field] = constraint
        
        return sheet_constraints


class ExcelFieldRecognizer:
    """Excel字段类型识别器"""
    
    TYPE_PATTERNS = {
        'name': {
            'keywords': ['姓名', '名字', 'name', '联系人', '用户姓名', '学员姓名', '员工姓名'],
            'generator': 'chinese_name'
        },
        'phone': {
            'keywords': ['手机', '电话', 'phone', 'tel', '联系方式', '移动电话', '手机号码'],
            'generator': 'chinese_phone'
        },
        'email': {
            'keywords': ['邮箱', '邮件', 'email', 'e-mail', '电子邮箱'],
            'generator': 'email'
        },
        'id_card': {
            'keywords': ['身份证', 'id_card', '身份证号', '证件号'],
            'generator': 'id_card'
        },
        'address': {
            'keywords': ['地址', 'address', '住址', '家庭地址', '居住地址'],
            'generator': 'chinese_address'
        },
        'province_city_district': {
            'keywords': ['省', '市', '区', '县', '省市', '市区', '区县', '省市区', '省市区县', '行政区划'],
            'generator': 'province_city_district'
        },
        'company': {
            'keywords': ['公司', '单位', 'company', '企业', '工作单位'],
            'generator': 'company_name'
        },
        'bank_card': {
            'keywords': ['银行卡', 'bank_card', '卡号'],
            'generator': 'bank_card'
        },
        'number': {
            'keywords': ['数值', '数字', 'number', '数量', '年龄'],
            'generator': 'random_int'
        },
        'date': {
            'keywords': ['日期', '时间', 'date', '年月日', '出生日期'],
            'generator': 'date'
        },
        'select': {
            'keywords': ['下拉框', '选择', '选项', 'select'],
            'generator': 'select'
        },
        'text': {
            'keywords': ['文本', '文字', 'text', '描述', '备注'],
            'generator': 'random_text'
        },
        'id': {
            'keywords': ['id', 'ID', '编号', '标识', 'uuid'],
            'generator': 'uuid'
        },
        'education': {
            'keywords': ['学历', '学位'],
            'generator': 'education_level'
        },
        'school': {
            'keywords': ['学校', '院校', '毕业院校'],
            'generator': 'school_name'
        },
        'subject': {
            'keywords': ['学科', '科目', '专业'],
            'generator': 'subject'
        },
        'ethnicity': {
            'keywords': ['民族'],
            'generator': 'ethnicity'
        },
        'gender': {
            'keywords': ['性别'],
            'generator': 'gender'
        },
    }
    
    @classmethod
    def recognize(cls, field_name: str) -> str:
        text = field_name.lower()
        for field_type, config in cls.TYPE_PATTERNS.items():
            for keyword in config['keywords']:
                if keyword.lower() in text:
                    return config['generator']
        return 'random_text'


class ExcelDataGenerator:
    """Excel数据生成器 - 使用第三方库生成数据"""
    
    @staticmethod
    def _generate_chinese_name():
        """使用faker生成中文姓名"""
        if HAS_FAKER and fake:
            return fake.name()
        # 备用方案
        surnames = ['王', '李', '张', '刘', '陈', '杨', '黄', '赵', '吴', '周']
        names = ['伟', '芳', '娜', '敏', '静', '丽', '强', '磊', '军', '洋']
        return random.choice(surnames) + random.choice(names)
    
    @staticmethod
    def _generate_phone():
        """使用faker生成手机号"""
        if HAS_FAKER and fake:
            phone = fake.phone_number()
            return phone.replace(' ', '').replace('-', '')
        # 备用方案
        prefixes = ['138', '139', '135', '136', '137', '150', '151', '152', '157', '158']
        return random.choice(prefixes) + ''.join(random.choices(string.digits, k=8))
    
    @staticmethod
    def _generate_email():
        """使用faker生成邮箱"""
        if HAS_FAKER and fake:
            return fake.email()
        # 备用方案
        domains = ['qq.com', '163.com', '126.com', 'gmail.com', 'sina.com']
        username = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        return f"{username}@{random.choice(domains)}"
    
    @staticmethod
    def _generate_id_card():
        """使用faker生成身份证号"""
        if HAS_FAKER and fake:
            return fake.ssn()
        # 备用方案
        areas = ['110101', '310101', '440106', '500103']
        area = random.choice(areas)
        year = random.randint(1960, 2005)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        birth = f"{year}{month:02d}{day:02d}"
        seq = ''.join(random.choices(string.digits, k=3))
        check = random.choice('0123456789X')
        return area + birth + seq + check
    
    @staticmethod
    def _generate_address():
        """使用faker生成地址"""
        if HAS_FAKER and fake:
            return fake.address()
        # 备用方案
        cities = ['北京市', '上海市', '广州市', '深圳市']
        districts = ['朝阳区', '海淀区', '浦东新区', '天河区']
        streets = ['建设大街', '解放路', '人民路', '中山路']
        return f"{random.choice(cities)}{random.choice(districts)}{random.choice(streets)}{random.randint(1, 999)}号"

    @staticmethod
    def _generate_province_city_district():
        """使用jionlp生成省市区县数据"""
        if HAS_JIONLP:
            try:
                location_data = jionlp.china_location_loader()
                
                # 获取所有省份（排除元数据键）
                provinces = [k for k in location_data.keys() if not k.startswith('_')]
                province = random.choice(provinces)
                
                # 获取该省下的城市
                city_data = location_data[province]
                cities = [k for k in city_data.keys() if not k.startswith('_')]
                
                if not cities:
                    return province
                
                city = random.choice(cities)
                
                # 获取该城市下的区县
                district_data = city_data[city]
                districts = [k for k in district_data.keys() if not k.startswith('_')]
                
                if not districts:
                    return f"{province}/{city}"
                
                district = random.choice(districts)
                
                return f"{province}/{city}/{district}"
            except Exception as e:
                # 如果出错，使用备用方案
                pass
        
        # 备用方案：使用faker生成省份+城市+区县的组合
        if HAS_FAKER and fake:
            return f"{fake.province()}/{fake.city()}/{fake.district() if hasattr(fake, 'district') else '区'}"
        
        # 最后的备用方案
        return "北京市/北京市/朝阳区"

    @staticmethod
    def _generate_company():
        """使用faker生成公司名称"""
        if HAS_FAKER and fake:
            return fake.company()
        # 备用方案
        cities = ['北京', '上海', '广州', '深圳']
        industries = ['科技', '信息', '网络', '软件', '教育']
        types = ['有限公司', '股份有限公司', '科技有限公司']
        return f"{random.choice(cities)}{random.choice(industries)}{random.choice(types)}"
    
    @staticmethod
    def _generate_bank_card():
        """使用faker生成银行卡号"""
        if HAS_FAKER and fake:
            # 生成16-19位银行卡号
            return fake.credit_card_number(card_type=None)
        # 备用方案
        return ''.join(random.choices(string.digits, k=random.randint(16, 19)))
    
    @staticmethod
    def _generate_date():
        """使用faker生成日期"""
        if HAS_FAKER and fake:
            return fake.date_between(start_date='-10y', end_date='today').strftime('%Y-%m-%d')
        # 备用方案
        return (datetime(2020, 1, 1) + timedelta(days=random.randint(0, 1825))).strftime('%Y-%m-%d')
    
    @classmethod
    def generate(cls, field_type: str, constraint: Dict = None) -> Any:
        """
        根据字段类型和约束生成数据
        
        Args:
            field_type: 字段类型
            constraint: 约束字典，包含 required, max_length, options 等
        """
        constraint = constraint or {}
        
        # 如果有预定义选项，优先从选项中选择
        if constraint.get('options'):
            return random.choice(constraint['options'])
        
        # 根据字段类型生成基础数据
        generators = {
            'chinese_name': cls._generate_chinese_name,
            'chinese_phone': cls._generate_phone,
            'email': cls._generate_email,
            'id_card': cls._generate_id_card,
            'chinese_address': cls._generate_address,
            'province_city_district': cls._generate_province_city_district,
            'company_name': cls._generate_company,
            'bank_card': cls._generate_bank_card,
            'random_int': lambda: random.randint(1, 100),
            'date': cls._generate_date,
            'select': lambda: random.choice(['选项A', '选项B', '选项C']),
            'random_text': lambda: ''.join(random.choices(string.ascii_letters + string.digits, k=10)),
            'uuid': lambda: str(RandomTools.random_uuid(count=1)['result']),
            'education_level': lambda: random.choice(['本科', '硕士研究生', '博士研究生', '大专', '高中']),
            'school_name': lambda: fake.company() + '大学' if HAS_FAKER and fake else '北京大学',
            'subject': lambda: random.choice(['语文', '数学', '英语', '音乐', '美术', '体育', '物理', '化学']),
            'ethnicity': lambda: random.choice(['汉族', '满族', '蒙古族', '回族', '藏族', '维吾尔族', '苗族']),
            'gender': lambda: random.choice(['男', '女']),
        }
        
        generator = generators.get(field_type, generators['random_text'])
        try:
            value = generator()
        except:
            value = f"测试_{field_type}_{random.randint(1000, 9999)}"
        
        # 应用长度限制
        max_length = constraint.get('max_length')
        if max_length and isinstance(value, str) and len(value) > max_length:
            value = value[:max_length]
        
        return value
    
    @classmethod
    def generate_with_constraint(cls, field_name: str, field_type: str, constraint: Dict = None) -> Any:
        """
        根据字段名、类型和约束智能生成数据
        
        优先使用约束中的选项，其次根据字段类型生成
        """
        constraint = constraint or {}
        
        # 如果有预定义选项，直接返回选项之一
        if constraint.get('options'):
            return random.choice(constraint['options'])
        
        # 否则根据字段类型生成
        return cls.generate(field_type, constraint)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_excel_template(request):
    """
    分析Excel模板字段
    支持单Sheet和多Sheet模板
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # 获取所有Sheet名称
        try:
            xl_file = pd.ExcelFile(excel_file)
            sheet_names = xl_file.sheet_names
        except Exception as e:
            return Response({'error': f'无法读取Excel文件: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not sheet_names:
            return Response({'error': 'Excel文件没有Sheet'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 读取第一个Sheet的原始数据（用于检查是否有填写须知）
        excel_file.seek(0)
        first_sheet_raw = pd.read_excel(excel_file, sheet_name=sheet_names[0], header=None)
        
        # 判断第一个Sheet是否有填写须知
        first_row = str(first_sheet_raw.iloc[0, 0]) if len(first_sheet_raw) > 0 else ''
        has_instruction_row = '填写须知' in first_row or len(first_row) > 100
        
        # 解析模板说明中的字段约束（多Sheet格式）
        sheet_constraints = {}
        if has_instruction_row:
            instruction_text = str(first_sheet_raw.iloc[0, 0])
            sheet_constraints = TemplateInstructionParser.parse_multi_sheet(instruction_text, sheet_names[0])
        
        # 分析所有Sheet
        sheets_info = []
        for idx, sheet_name in enumerate(sheet_names):
            excel_file.seek(0)
            
            # 所有Sheet都使用header=1，因为每个Sheet第一行都是填写须知
            df = pd.read_excel(excel_file, sheet_name=idx, header=1)
            
            # 分析该Sheet的字段
            fields = []
            for col in df.columns:
                field_name = str(col)
                field_type = ExcelFieldRecognizer.recognize(field_name)
                
                # 获取字段约束（从对应Sheet的约束中查找）
                constraints_for_sheet = sheet_constraints.get(sheet_name, {})
                constraint = constraints_for_sheet.get(field_name, {})
                
                # 如果模板说明中有选项，更新字段类型为select
                if constraint.get('options'):
                    field_type = 'select'
                
                fields.append({
                    'name': field_name,
                    'type': field_type,
                    'nullable': not constraint.get('required', False),
                    'constraint': constraint
                })
            
            sheets_info.append({
                'name': sheet_name,
                'fields': fields,
                'total_fields': len(fields),
                'data_rows': len(df)
            })
        
        return Response({
            'success': True,
            'sheets': sheets_info,
            'total_sheets': len(sheets_info),
            'has_instruction': has_instruction_row,
            # 兼容单Sheet的旧格式
            'fields': sheets_info[0]['fields'] if sheets_info else [],
            'total_fields': sheets_info[0]['total_fields'] if sheets_info else 0,
            'data_rows': sheets_info[0]['data_rows'] if sheets_info else 0
        })
        
    except Exception as e:
        return Response({'error': f'分析失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _fill_single_sheet(ws, headers, row_count, custom_fields, constraints, has_instruction_row, instruction_text=None):
    """
    填充单个Sheet的数据
    """
    current_row = 1
    
    # 第1行：说明行（如果有）- 合并所有单元格
    if has_instruction_row and instruction_text:
        ws.cell(row=current_row, column=1, value=instruction_text)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        instruction_cell = ws.cell(row=current_row, column=1)
        instruction_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        current_row += 1
    
    # 字段名行 - 添加绿色背景
    header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    header_row = current_row
    current_row += 1
    
    # 生成数据
    for i in range(row_count):
        for col_idx, header in enumerate(headers, 1):
            header_str = str(header)
            if header_str in custom_fields:
                value = custom_fields[header_str]
            else:
                field_type = ExcelFieldRecognizer.recognize(header_str)
                constraint = constraints.get(header_str, {})
                value = ExcelDataGenerator.generate_with_constraint(header_str, field_type, constraint)
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
    
    # 设置列宽自适应
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row_idx in range(header_row + 1, current_row):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, 15), 50)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    return current_row


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fill_excel_data(request):
    """
    填充Excel数据
    支持单Sheet和多Sheet模板
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        row_count = int(request.data.get('row_count', 10))

        # 获取自定义字段值
        custom_fields = {}
        if 'custom_fields' in request.data:
            try:
                custom_fields = json.loads(request.data['custom_fields'])
            except:
                pass
        
        # 获取所有Sheet名称
        try:
            xl_file = pd.ExcelFile(excel_file)
            sheet_names = xl_file.sheet_names
        except Exception as e:
            return Response({'error': f'无法读取Excel文件: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not sheet_names:
            return Response({'error': 'Excel文件没有Sheet'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 读取第一个Sheet的原始数据（用于检查是否有填写须知）
        excel_file.seek(0)
        first_sheet_raw = pd.read_excel(excel_file, sheet_name=sheet_names[0], header=None)
        
        # 判断第一个Sheet是否有填写须知
        first_row = str(first_sheet_raw.iloc[0, 0]) if len(first_sheet_raw) > 0 else ''
        has_instruction_row = '填写须知' in first_row or len(first_row) > 100
        
        # 解析模板说明中的字段约束（多Sheet格式）
        sheet_constraints = {}
        instruction_text = None
        if has_instruction_row:
            instruction_text = str(first_sheet_raw.iloc[0, 0])
            sheet_constraints = TemplateInstructionParser.parse_multi_sheet(instruction_text, sheet_names[0])
        
        # 使用openpyxl创建Excel文件
        wb = Workbook()
        
        # 处理所有Sheet
        for idx, sheet_name in enumerate(sheet_names):
            excel_file.seek(0)
            
            # 读取当前Sheet的原始数据（获取填写须知）
            sheet_raw = pd.read_excel(excel_file, sheet_name=idx, header=None)
            sheet_first_row = str(sheet_raw.iloc[0, 0]) if len(sheet_raw) > 0 else ''
            sheet_has_instruction = '填写须知' in sheet_first_row or len(sheet_first_row) > 100
            sheet_instruction_text = str(sheet_raw.iloc[0, 0]) if sheet_has_instruction else None
            
            # 所有Sheet都使用header=1，因为每个Sheet第一行都是填写须知
            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name=idx, header=1)
            
            headers = df.columns.tolist()
            
            # 获取或创建工作表
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # 获取该Sheet的约束
            constraints_for_sheet = sheet_constraints.get(sheet_name, {})
            
            # 填充Sheet数据 - 每个Sheet都显示自己的填写须知
            _fill_single_sheet(
                ws, headers, row_count, custom_fields, 
                constraints_for_sheet, sheet_has_instruction, 
                sheet_instruction_text
            )
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        original_filename = excel_file.name
        if '模板' in original_filename:
            base_name = original_filename.split('模板')[0].rstrip('-_')
        else:
            base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename

        username = request.user.username if hasattr(request, 'user') and request.user else 'unknown'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{base_name}_{username}_{timestamp}.xlsx"

        from urllib.parse import quote
        filename_utf8 = quote(filename.encode('utf-8'))

        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename_utf8}"
        
        return response
        
    except Exception as e:
        return Response({'error': f'填充失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def preview_filled_data(request):
    """
    预览填充数据（不生成文件）
    支持单Sheet和多Sheet模板
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        row_count = min(int(request.data.get('row_count', 5)), 20)  # 最多预览20条

        # 获取自定义字段值
        custom_fields = {}
        if 'custom_fields' in request.data:
            try:
                custom_fields = json.loads(request.data['custom_fields'])
            except:
                pass
        
        # 获取所有Sheet名称
        try:
            xl_file = pd.ExcelFile(excel_file)
            sheet_names = xl_file.sheet_names
        except Exception as e:
            return Response({'error': f'无法读取Excel文件: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not sheet_names:
            return Response({'error': 'Excel文件没有Sheet'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 读取第一个Sheet的原始数据（用于检查是否有填写须知）
        excel_file.seek(0)
        first_sheet_raw = pd.read_excel(excel_file, sheet_name=sheet_names[0], header=None)
        
        # 判断第一个Sheet是否有填写须知
        first_row = str(first_sheet_raw.iloc[0, 0]) if len(first_sheet_raw) > 0 else ''
        has_instruction_row = '填写须知' in first_row or len(first_row) > 100
        
        # 解析模板说明中的字段约束（多Sheet格式）
        sheet_constraints = {}
        if has_instruction_row:
            instruction_text = str(first_sheet_raw.iloc[0, 0])
            sheet_constraints = TemplateInstructionParser.parse_multi_sheet(instruction_text, sheet_names[0])
        
        # 预览所有Sheet的数据
        all_sheets_preview = []
        all_sheets_fields = []
        
        for idx, sheet_name in enumerate(sheet_names):
            excel_file.seek(0)
            
            # 所有Sheet都使用header=1，因为每个Sheet第一行都是填写须知
            df = pd.read_excel(excel_file, sheet_name=idx, header=1)
            
            # 获取该Sheet的约束
            constraints_for_sheet = sheet_constraints.get(sheet_name, {})
            
            # 生成该Sheet的预览数据
            sheet_preview_data = []
            for i in range(row_count):
                row = {}
                for col in df.columns:
                    col_str = str(col)
                    # 如果该字段有自定义值，则使用自定义值
                    if col_str in custom_fields:
                        row[col_str] = custom_fields[col_str]
                    else:
                        field_type = ExcelFieldRecognizer.recognize(col_str)
                        constraint = constraints_for_sheet.get(col_str, {})
                        row[col_str] = ExcelDataGenerator.generate_with_constraint(col_str, field_type, constraint)
                sheet_preview_data.append(row)
            
            # 字段分析
            sheet_fields = []
            for col in df.columns:
                field_name = str(col)
                field_type = ExcelFieldRecognizer.recognize(field_name)
                constraint = constraints_for_sheet.get(field_name, {})
                
                # 如果模板说明中有选项，更新字段类型为select
                if constraint.get('options'):
                    field_type = 'select'
                
                sheet_fields.append({
                    'name': field_name,
                    'type': field_type,
                    'constraint': constraint
                })
            
            all_sheets_preview.append({
                'sheet_name': sheet_name,
                'preview_data': sheet_preview_data,
                'fields': sheet_fields,
                'total_fields': len(sheet_fields)
            })
            
            # 兼容旧格式，保留第一个Sheet的数据
            if idx == 0:
                all_sheets_fields = sheet_fields
        
        return Response({
            'success': True,
            'fields': all_sheets_fields,
            'preview_data': all_sheets_preview[0]['preview_data'] if all_sheets_preview else [],
            'sheets_preview': all_sheets_preview,
            'total_fields': len(all_sheets_fields),
            'has_instruction': has_instruction_row,
            'total_sheets': len(sheet_names),
            'current_sheet': sheet_names[0]
        })
        
    except Exception as e:
        return Response({'error': f'预览失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
